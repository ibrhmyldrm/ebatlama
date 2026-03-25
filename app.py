from flask import Flask, render_template, request, jsonify, send_file
import json, io, os, threading, sys, webview
from datetime import datetime

app = Flask(__name__)

# ─── Yardımcı: Windows dosya dialog'ları (thread-safe) ───────────────────────
def _run_in_main_thread(fn):
    """Flask worker thread'inden tkinter'ı ana thread'de çalıştır."""
    result = [None]
    done   = threading.Event()
    def wrapper():
        result[0] = fn()
        done.set()
    # Ana thread'e gönder — tkinter sadece ana thread'de çalışır
    import tkinter as tk
    # Doğrudan çalıştır (Flask thread'i de çalışabiliyor Windows'ta)
    result[0] = fn()
    return result[0]

def ask_save_path(default_name, filetypes):
    """Windows Farklı Kaydet dialog'u."""
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.lift()
        root.attributes('-topmost', True)
        root.focus_force()
        path = filedialog.asksaveasfilename(
            parent=root,
            initialfile=default_name,
            defaultextension=filetypes[0][1].replace("*", ""),
            filetypes=filetypes,
            title="Kayit yeri sec"
        )
        root.destroy()
        return path if path else None
    except Exception as e:
        import traceback; traceback.print_exc()
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        docs    = os.path.join(os.path.expanduser("~"), "Documents")
        folder  = desktop if os.path.exists(desktop) else (docs if os.path.exists(docs) else os.path.expanduser("~"))
        return os.path.join(folder, default_name)

def ask_string(title, prompt_text, initial=""):
    """Tkinter ile proje adı soran dialog."""
    try:
        import tkinter as tk
        from tkinter import simpledialog
        root = tk.Tk()
        root.withdraw()
        root.lift()
        root.attributes('-topmost', True)
        root.focus_force()
        result = simpledialog.askstring(title, prompt_text, initialvalue=initial, parent=root)
        root.destroy()
        return result
    except Exception:
        return initial or "proje"

# ─── Maximal Rectangles algoritması (endüstri standardı) ──────────────────
def optimize_cuts(pieces, sheet_w, sheet_h, blade=4):
    """
    Maximal Rectangles + çoklu strateji + çoklu skor fonksiyonu.
    Testere uyumlu: yatay/dikey kesimler, guillotine kısıtı.
    """
    import random

    def intersects(ax,ay,aw,ah,bx,by,bw,bh):
        return not (ax+aw<=bx or bx+bw<=ax or ay+ah<=by or by+bh<=ay)

    def split_free(free_rects, px, py, pw, ph):
        new_free=[]; bpw,bph=pw+blade,ph+blade
        for rx,ry,rw,rh in free_rects:
            if not intersects(rx,ry,rw,rh,px,py,bpw,bph):
                new_free.append((rx,ry,rw,rh)); continue
            if px>rx:          new_free.append((rx,ry,px-rx,rh))
            if px+bpw<rx+rw:   new_free.append((px+bpw,ry,(rx+rw)-(px+bpw),rh))
            if py>ry:          new_free.append((rx,ry,rw,py-ry))
            if py+bph<ry+rh:   new_free.append((rx,py+bph,rw,(ry+rh)-(py+bph)))
        pruned=[]
        for i,r in enumerate(new_free):
            if r[2]<=0 or r[3]<=0: continue
            inside=any(
                j!=i and s[0]<=r[0] and s[1]<=r[1] and
                s[0]+s[2]>=r[0]+r[2] and s[1]+s[3]>=r[1]+r[3]
                for j,s in enumerate(new_free)
            )
            if not inside: pruned.append(r)
        return pruned

    def pack_sheet(pieces_list, score_fn):
        free=[(0,0,sheet_w,sheet_h)]; placed=[]; remaining=pieces_list[:]
        while remaining:
            best=None; best_score=float('inf'); best_i=-1
            for i,(pw,ph,label,bant) in enumerate(remaining):
                for rx,ry,rw,rh in free:
                    for fpw,fph,rot in [(pw,ph,False),(ph,pw,True)]:
                        if fpw<=rw and fph<=rh:
                            score=score_fn(rw,rh,fpw,fph)
                            if score<best_score:
                                best_score=score
                                best=(rx,ry,fpw,fph,label,bant,rot); best_i=i
            if best is None: break
            rx,ry,fpw,fph,label,bant,rot=best
            placed.append((rx,ry,fpw,fph,label,bant,rot))
            remaining.pop(best_i)
            free=split_free(free,rx,ry,fpw,fph)
        return placed,remaining,free

    # Skor fonksiyonları
    SCORES = [
        lambda rw,rh,pw,ph: min(rw-pw, rh-ph),          # BSSF
        lambda rw,rh,pw,ph: max(rw-pw, rh-ph),           # BLSF
        lambda rw,rh,pw,ph: rw*rh - pw*ph,               # BAF
        lambda rw,rh,pw,ph: -(rw*rh - pw*ph),            # WAF
    ]

    expanded=[]
    for w,h,label,qty,bant in pieces:
        for _ in range(qty): expanded.append((w,h,label,bant))

    by_name={}
    for p in expanded: by_name.setdefault(p[2],[]).append(p)

    def make_grouped(key_fn,rev=True):
        order=[]
        for name in sorted(by_name.keys(),key=key_fn,reverse=rev):
            order.extend(by_name[name])
        return order

    base_orders=[
        sorted(expanded,key=lambda p:max(p[0],p[1]),reverse=True),
        sorted(expanded,key=lambda p:p[0]*p[1],reverse=True),
        sorted(expanded,key=lambda p:min(p[0],p[1]),reverse=True),
        sorted(expanded,key=lambda p:p[0]+p[1],reverse=True),
        make_grouped(lambda n:max(by_name[n][0][0],by_name[n][0][1])),
        make_grouped(lambda n:by_name[n][0][0]*by_name[n][0][1]),
        make_grouped(lambda n:len(by_name[n])*by_name[n][0][0]*by_name[n][0][1]),
    ]
    random.seed(42)
    rand_orders=[]
    for _ in range(30):
        s=expanded[:]; random.shuffle(s); rand_orders.append(s)

    best_sheets=float('inf'); best_waste=float('inf')
    best_fires=float('inf'); best_pl=[]; best_free=[]

    for order in base_orders+rand_orders:
        for score_fn in SCORES:
            all_pl=[]; all_free=[]; remaining=order[:]; idx=0
            while remaining:
                pl,remaining,free=pack_sheet(remaining,score_fn)
                if not pl: break
                for p in pl: all_pl.append((idx,)+p)
                all_free.append(free); idx+=1
            if remaining: continue
            total=sheet_w*sheet_h*idx; used=sum(p[3]*p[4] for p in all_pl)
            waste=round((1-used/total)*100,1) if total else 0
            fires=sum(len(f) for f in all_free)
            if (idx<best_sheets or
                (idx==best_sheets and fires<best_fires) or
                (idx==best_sheets and fires==best_fires and waste<best_waste)):
                best_sheets=idx; best_waste=waste; best_fires=fires
                best_pl=all_pl; best_free=all_free

    return best_sheets,best_pl,best_waste,best_free

# ─── Maliyet Hesaplama ────────────────────────────────────────────────────────
def calculate_cost(data):
    pieces     = data.get("pieces", [])
    pricing    = data.get("pricing", {})
    sheet_w    = float(data.get("sheet_w", 2440))
    sheet_h    = float(data.get("sheet_h", 1220))
    sheet_price = float(pricing.get("sheet_price", 0))
    cut_price   = float(pricing.get("cut_price",   0))
    band_price  = float(pricing.get("band_price",  0))
    shipping    = float(pricing.get("shipping",    0))

    piece_list, total_band_m = [], 0.0

    for p in pieces:
        w_cm = float(p.get("w", 0))   # EN  (kısa kenar = KK)
        h_cm = float(p.get("h", 0))   # BOY (uzun kenar = UK)
        w_mm = w_cm * 10
        h_mm = h_cm * 10
        qty  = int(p.get("qty", 1))
        bant = p.get("bant", [])
        piece_list.append((w_mm, h_mm, p.get("name", "Parça"), qty, bant))

        # UK-1/UK-2 = uzun kenar = BOY (h)
        # KK-1/KK-2 = kısa kenar = EN  (w)
        for side in bant:
            if side in ["ust", "alt"]:      # UK → BOY boyunca
                total_band_m += (h_mm / 1000) * qty
            elif side in ["sol", "sag"]:    # KK → EN boyunca
                total_band_m += (w_mm / 1000) * qty

    blade = float(data.get("blade", 3))
    num_sheets, placements, waste_pct, sheets_free = optimize_cuts(piece_list, sheet_w, sheet_h, blade=blade)

    cuts_per = {}
    for p in placements:
        cuts_per[p[0]] = cuts_per.get(p[0], 0) + 1
    total_cuts = sum(max(1, v) for v in cuts_per.values())

    cost_sheets = num_sheets * sheet_price
    cost_cuts   = num_sheets * cut_price
    cost_band   = total_band_m * band_price
    cost_total  = cost_sheets + cost_cuts + cost_band + shipping

    # Fire alanları — her levha için kalan boş dikdörtgenler
    waste_rects = []
    for si, free_list in enumerate(sheets_free):
        for rx, ry, rw, rh in free_list:
            if rw > 0 and rh > 0:
                waste_rects.append({
                    "sheet": si + 1,
                    "x": rx, "y": ry,
                    "w": rw, "h": rh,
                    "w_cm": round(rw/10, 1),
                    "h_cm": round(rh/10, 1),
                })

    return {
        "num_sheets":    num_sheets,
        "waste_pct":     waste_pct,
        "total_cuts":    total_cuts,
        "total_band_m":  round(total_band_m, 2),
        "cost_sheets":   round(cost_sheets,  2),
        "cost_cuts":     round(cost_cuts,    2),
        "cost_band":     round(cost_band,    2),
        "cost_shipping": round(shipping,     2),
        "cost_total":    round(cost_total,   2),
        "waste_rects":   waste_rects,
        "placements": [
            {
                "sheet":    p[0]+1,
                "x": p[1], "y": p[2],
                "w": p[3], "h": p[4],
                "label":    p[5],
                "bant":     p[6],
                "rotated":  p[7],
                "bant_visual": _calc_bant_visual(p[6], p[7], p[3], p[4]),
                "w_cm":     round(p[3]/10, 1),
                "h_cm":     round(p[4]/10, 1),
            }
            for p in placements
        ],
        "sheet_w": sheet_w, "sheet_h": sheet_h,
    }

def _calc_bant_visual(bant, rotated, pw, ph):
    """
    Placement'taki gerçek w(pw) ve h(ph) değerlerine bakarak
    UK ve KK'nın hangi yönde olduğunu belirle.
    
    UK = uzun kenar:
      - ph >= pw ise → uzun kenar dikey → UK = sol/sag
      - pw > ph  ise → uzun kenar yatay → UK = ust/alt
    
    bant listesindeki değerler:
      ust/alt = UK-1/UK-2 (uzun kenar)
      sol/sag = KK-1/KK-2 (kısa kenar)
    
    Bunları gerçek w/h'ye göre yeniden eşleştir.
    """
    visual = []
    uk_is_vertical = (ph >= pw)   # uzun kenar dikey mi?

    for side in bant:
        if side == "ust":    # UK-1
            visual.append("sol" if uk_is_vertical else "ust")
        elif side == "alt":  # UK-2
            visual.append("sag" if uk_is_vertical else "alt")
        elif side == "sol":  # KK-1
            visual.append("ust" if uk_is_vertical else "sol")
        elif side == "sag":  # KK-2
            visual.append("alt" if uk_is_vertical else "sag")
    return visual


# ─── Routes ───────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/calculate", methods=["POST"])
def api_calculate():
    return jsonify(calculate_cost(request.json))


# ─── Yardımcı: Levha canvas görüntüsü oluştur (PIL) ─────────────────────────
def draw_sheet_image(placements, sheet_w, sheet_h, waste_rects=None,
                     piece_labels=None, img_w=2400, title=""):
    from PIL import Image, ImageDraw, ImageFont
    import io as _io

    COLORS_LIST = [
        "#A8D5BA","#F7C59F","#B5C7E3","#F2A7BB","#C5B4E3",
        "#A8C7D5","#EDD5A3","#B4D1B0","#D5B4A8","#C7D5A8",
        "#9DC4E0","#F0C9A0","#B8E0C8","#E8B4D0","#C4D4A8",
    ]
    pnames    = list({p["label"].replace(" (D)","") for p in placements})
    color_map = {n: COLORS_LIST[i % len(COLORS_LIST)] for i, n in enumerate(pnames)}

    PAD    = 60
    TITLE  = 50
    scale  = (img_w - 2*PAD) / sheet_w
    img_h  = int(sheet_h * scale) + 2*PAD + TITLE
    img    = Image.new("RGB", (img_w, img_h), "#FFFFFF")
    draw   = ImageDraw.Draw(img)

    font_size_title = max(32, int(img_w / 55))
    font_size_label = max(26, int(img_w / 70))
    font_size_num   = max(22, int(img_w / 85))
    font_size_dim   = max(20, int(img_w / 95))

    def load_font(size):
        candidates = [
            os.path.join(os.environ.get("WINDIR","C:\\Windows"), "Fonts", "arial.ttf"),
            os.path.join(os.environ.get("WINDIR","C:\\Windows"), "Fonts", "calibri.ttf"),
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
            os.path.join(_get_static_dir(), "DejaVuSans.ttf"),
        ]
        for p in candidates:
            if os.path.exists(p):
                try: return ImageFont.truetype(p, size)
                except: pass
        return ImageFont.load_default()

    ft = load_font(font_size_title)
    fl = load_font(font_size_label)
    fn = load_font(font_size_num)
    fd = load_font(font_size_dim)

    # Başlık
    draw.text((PAD, 8), title, font=ft, fill="#4A1A8A")
    dim_txt = f"{int(sheet_w)} x {int(sheet_h)} mm"
    draw.text((img_w - PAD - 400, 8), dim_txt, font=fd, fill="#7B5EA7")

    # Levha çerçevesi
    lx1 = PAD
    ly1 = PAD + TITLE
    lx2 = PAD + int(sheet_w * scale)
    ly2 = PAD + TITLE + int(sheet_h * scale)
    draw.rectangle([lx1, ly1, lx2, ly2], fill="#F0EBF8", outline="#4A1A8A", width=4)

    # ── Fire alanları ──
    if waste_rects:
        import itertools
        # Diagonal çizgi deseni
        for i, wr in enumerate(waste_rects):
            wx1 = lx1 + int(wr["x"] * scale)
            wy1 = ly1 + int(wr["y"] * scale)
            wx2 = wx1 + int(wr["w"] * scale)
            wy2 = wy1 + int(wr["h"] * scale)
            if wx2 <= wx1 or wy2 <= wy1: continue

            # Açık sarı arka plan
            draw.rectangle([wx1, wy1, wx2, wy2], fill="#FAF6E8", outline="#C8A050", width=2)

            # Çapraz çizgiler
            step = max(20, int(scale * 15))
            for offset in range(0, (wx2-wx1)+(wy2-wy1), step):
                x_s = wx1 + offset; y_s = wy1
                x_e = wx1;          y_e = wy1 + offset
                draw.line([max(wx1,x_s), max(wy1,y_s), max(wx1,x_e), max(wy1,y_e)],
                          fill="#D4B870", width=1)

            # Fire numarası ve ölçü
            fcx = (wx1 + wx2) // 2
            fcy = (wy1 + wy2) // 2
            fw = wx2 - wx1; fh = wy2 - wy1
            if fw > 80 and fh > 50:
                fire_txt = f"FİRE {i+1}"
                try:
                    tb = draw.textbbox((0,0), fire_txt, font=fd)
                    tw = tb[2]-tb[0]
                    draw.text((fcx - tw//2, fcy - font_size_dim), fire_txt, font=fd, fill="#A08040")
                    dim_t = f"{wr['w_cm']}x{wr['h_cm']}"
                    tb2 = draw.textbbox((0,0), dim_t, font=fd)
                    tw2 = tb2[2]-tb2[0]
                    draw.text((fcx - tw2//2, fcy + 4), dim_t, font=fd, fill="#A08040")
                except:
                    draw.text((fcx, fcy), f"FIRE {i+1}", font=fd, fill="#A08040")

    bw = max(4, int(scale * 5))

    # ── Parçalar ──
    for pp in placements:
        x1 = lx1 + int(pp["x"] * scale)
        y1 = ly1 + int(pp["y"] * scale)
        x2 = x1  + int(pp["w"] * scale)
        y2 = y1  + int(pp["h"] * scale)
        fc = color_map.get(pp["label"].replace(" (D)",""), "#C5B4E3")

        draw.rectangle([x1, y1, x2, y2], fill=fc, outline="#4A1A8A", width=2)

        # Bant kenarları
        bant_sides = pp.get("bant_visual") or pp.get("bant") or []
        for side in bant_sides:
            if   side == "ust": draw.line([x1+2, y1, x2-2, y1], fill="#E07020", width=bw)
            elif side == "alt": draw.line([x1+2, y2, x2-2, y2], fill="#E07020", width=bw)
            elif side == "sol": draw.line([x1, y1+2, x1, y2-2], fill="#E07020", width=bw)
            elif side == "sag": draw.line([x2, y1+2, x2, y2-2], fill="#E07020", width=bw)

        # Etiket
        pw = x2 - x1
        ph = y2 - y1
        base = pp["label"].replace(" (D)","")
        key  = f"{pp.get('sheet',1)}_{pp['x']}_{pp['y']}"
        plabel = piece_labels.get(key, "") if piece_labels else ""
        cx = x1 + pw // 2
        cy = y1 + ph // 2

        if pw > 60 and ph > 40:
            try:
                # Üst içte: EN ölçüsü yatay
                w_txt = f"{pp['w_cm']}"
                wb2 = draw.textbbox((0,0), w_txt, font=fd)
                wtw = wb2[2]-wb2[0]
                draw.text((cx - wtw//2, y1 + 6), w_txt, font=fd, fill="#3A1800")

                # Sol içte: BOY ölçüsü dikey
                h_txt = f"{pp['h_cm']}"
                from PIL import Image as _PImg
                tmp = _PImg.new("RGBA", (300, 50), (0,0,0,0))
                tdraw = ImageDraw.Draw(tmp)
                tdraw.text((0, 5), h_txt, font=fd, fill=(58,24,0,255))
                tmp_r = tmp.rotate(90, expand=True)
                hb = draw.textbbox((0,0), h_txt, font=fd)
                htw = hb[2]-hb[0]
                img.paste(tmp_r, (x1 + 6, cy - htw//2), tmp_r)

                # Orta: parça adı
                nb = draw.textbbox((0,0), base, font=fl)
                nw, nh = nb[2]-nb[0], nb[3]-nb[1]
                draw.text((cx - nw//2, cy - nh//2 - font_size_num//2), base, font=fl, fill="#1A0533")

                # Altında numara etiketi
                if plabel:
                    lb = draw.textbbox((0,0), plabel, font=fn)
                    lw2 = lb[2]-lb[0]
                    draw.text((cx - lw2//2, cy + nh//2), plabel, font=fn, fill="#4A1A8A")
            except Exception as e:
                draw.text((cx, cy), base, font=fl, fill="#1A0533")

    # Legend
    leg_y = img_h - 45
    draw.rectangle([PAD, leg_y-5, PAD+350, leg_y+35], fill="#F7F3FF", outline="#D4C4F0", width=1)
    draw.line([PAD+10, leg_y+15, PAD+60, leg_y+15], fill="#E07020", width=4)
    draw.text((PAD+70, leg_y+4), "= Bant kenari", font=fd, fill="#555555")

    buf = _io.BytesIO()
    img.save(buf, format="PNG", dpi=(150, 150))
    buf.seek(0)
    return buf


# ─── Excel Export ─────────────────────────────────────────────────────────────
@app.route("/api/export_excel", methods=["POST"])
def api_export_excel():
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.worksheet.page import PageMargins
    except ImportError:
        return jsonify({"error": "openpyxl yuklu degil"}), 500
    try:
        data    = request.json
        result  = calculate_cost(data)
        pieces  = data.get("pieces", [])
        pricing = data.get("pricing", {})

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kesim Listesi"

        # ── Print ayarları (A4 dikey) ──
        ws.page_setup.orientation        = "portrait"
        ws.page_setup.paperSize           = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage           = True
        ws.page_setup.fitToWidth          = 1
        ws.page_setup.fitToHeight         = 0
        ws.print_options.horizontalCentered = True
        ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75,
                                       header=0.3, footer=0.3)

        # ── Stiller ──
        H_FILL = PatternFill("solid", fgColor="4A1A8A")
        S_FILL = PatternFill("solid", fgColor="7B5EA7")
        A_FILL = PatternFill("solid", fgColor="D4C4F0")
        W_FILL = PatternFill("solid", fgColor="FFFFFF")
        T_FILL = PatternFill("solid", fgColor="1A0533")
        thin   = Side(style="thin", color="A66EE8")
        BRD    = Border(left=thin, right=thin, top=thin, bottom=thin)
        CTR    = Alignment(horizontal="center", vertical="center", wrap_text=False)
        LFT    = Alignment(horizontal="left",   vertical="center", wrap_text=False)

        def cell(ws, r, c, val="", fill=None, font=None, align=None, border=None, height=None):
            cl = ws.cell(row=r, column=c, value=val)
            if fill:   cl.fill   = fill
            if font:   cl.font   = font
            if align:  cl.alignment = align
            if border: cl.border = border
            if height: ws.row_dimensions[r].height = height
            return cl

        FH14 = Font(bold=True, color="FFFFFF", size=14)
        FH10 = Font(bold=True, color="EDE0FF", size=10)
        FH11 = Font(bold=True, color="FFFFFF", size=11)
        FN10 = Font(color="1A0533", size=10)
        FT11 = Font(bold=True, color="FFFFFF", size=11)

        # ── Sütun genişlikleri ──
        # A=#, B=ParcaAdi, C=Boy, D=En, E=Adet, F=BantKenarlari, G=Bant(m), H=Levha#
        col_ws = [5, 28, 10, 10, 7, 20, 10, 10]
        for i, w in enumerate(col_ws, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # ── Satır 1: Başlık (tüm sütunlara renk, sadece A'ya metin) ──
        baslik = f"ATOLYEMHANEM  -  KESIM LISTESI  -  {datetime.now().strftime('%d.%m.%Y')}"
        for c in range(1, 9):
            cl = ws.cell(row=1, column=c, value=(baslik if c==1 else ""))
            cl.fill = H_FILL
            cl.font = Font(bold=True, color="FFFFFF", size=12)
            cl.alignment = Alignment(horizontal="left" if c==1 else "center",
                                     vertical="center", wrap_text=False)
            cl.border = BRD
        ws.row_dimensions[1].height = 22

        # ── Satır 2: Levha bilgisi ──
        info = (f"Levha Adedi: {result['num_sheets']}  |  "
                f"Malzeme: {data.get('material','MDF')}  |  "
                f"Boyut: {int(data.get('sheet_w',2440))}x{int(data.get('sheet_h',1220))} mm  |  "
                f"Toplam Bant: {result['total_band_m']} m  |  "
                f"Fire: %{result['waste_pct']}")
        for c in range(1, 9):
            cl = ws.cell(row=2, column=c, value=(info if c==1 else ""))
            cl.fill = S_FILL
            cl.font = Font(bold=True, color="EDE0FF", size=9)
            cl.alignment = Alignment(horizontal="left" if c==1 else "center",
                                     vertical="center", wrap_text=False)
            cl.border = BRD
        ws.row_dimensions[2].height = 16

        # ── Satır 3: boş ──
        ws.row_dimensions[3].height = 6

        # ── Satır 4: Sütun başlıkları ──
        hdrs = ["#", "Parca Adi", "Boy(cm)", "En(cm)", "Adet", "Bant Kenarlari", "Bant(m)", "Levha#"]
        for c, h in enumerate(hdrs, 1):
            cell(ws, 4, c, val=h,
                 fill=PatternFill("solid", fgColor="2D0D5C"),
                 font=FH11, align=CTR, border=BRD, height=18)

        # ── Parça satırları ──
        pl_by_label = {}
        for p in result["placements"]:
            lbl = p["label"].replace(" (D)", "")
            pl_by_label.setdefault(lbl, []).append(str(p["sheet"]))

        bant_map = {"ust":"UK-1","alt":"UK-2","sol":"KK-1","sag":"KK-2"}
        COLORS_LIST = ["A8D5BA","F7C59F","B5C7E3","F2A7BB","C5B4E3",
                       "A8C7D5","EDD5A3","B4D1B0","D5B4A8","C7D5A8"]

        for idx, p in enumerate(pieces):
            w_cm = float(p.get("w", 0))
            h_cm = float(p.get("h", 0))
            qty  = int(p.get("qty", 1))
            bant = p.get("bant", [])
            bstr = ", ".join(bant_map.get(s, s) for s in bant) or "-"
            name = p.get("name", f"Parca {idx+1}")
            bm   = 0
            for side in bant:
                if side in ["ust","alt"]: bm += (h_cm*10/1000)*qty
                else:                     bm += (w_cm*10/1000)*qty
            sh_used = ", ".join(sorted(set(pl_by_label.get(name, []))))

            # Parçaya özel renk — yerleşim planı ile eşleşsin
            pc_fill = PatternFill("solid", fgColor=COLORS_LIST[idx % len(COLORS_LIST)])
            r = 5 + idx
            vals = [idx+1, name, h_cm, w_cm, qty, bstr, round(bm,2), sh_used]
            aligns = [CTR, LFT, CTR, CTR, CTR, CTR, CTR, CTR]
            for c, (val, aln) in enumerate(zip(vals, aligns), 1):
                cell(ws, r, c, val=val, fill=pc_fill, font=FN10, align=aln, border=BRD, height=16)

        # ── Boş satır ──
        sep = 5 + len(pieces)
        ws.row_dimensions[sep].height = 8

        # ── Maliyet özeti ──
        mrow = sep + 1
        for c in range(1, 9):
            cl = ws.cell(row=mrow, column=c,
                         value=("" if c!=2 else "MALIYET OZETI"))
            cl.fill = H_FILL
            cl.font = FH11
            cl.alignment = CTR
            cl.border = BRD
        ws.row_dimensions[mrow].height = 18

        cost_rows = [
            ("Levha Maliyeti", result["cost_sheets"],   f"{result['num_sheets']} levha x {pricing.get('sheet_price',0)} TL"),
            ("Kesim Ucreti",   result["cost_cuts"],     f"{result['num_sheets']} plaka x {pricing.get('cut_price',0)} TL"),
            ("PVC Bant",       result["cost_band"],     f"{result['total_band_m']} m x {pricing.get('band_price',0)} TL"),
            ("Nakliye",        result["cost_shipping"], ""),
            ("TOPLAM",         result["cost_total"],    ""),
        ]
        for i, (lbl, amt, note) in enumerate(cost_rows):
            r    = mrow + 1 + i
            is_t = lbl == "TOPLAM"
            fill = T_FILL if is_t else (A_FILL if i%2==0 else W_FILL)
            font = FT11   if is_t else FN10
            # B=etiket(geniş), C=tutar, D=detay notu
            row_vals = ["", lbl, f"{amt} TL", note, "", "", "", ""]
            row_aln  = [CTR, LFT, CTR, LFT, CTR, CTR, CTR, CTR]
            for c, (val, aln) in enumerate(zip(row_vals, row_aln), 1):
                cell(ws, r, c, val=val, fill=fill, font=font, align=aln, border=BRD, height=16)

        # ── Canvas: Tüm levhalar tek sayfada (2. sheet) ──
        try:
            from PIL import Image as PILImage
            ws2 = wb.create_sheet("Levha Plani")
            ws2.page_setup.orientation = "portrait"
            ws2.page_setup.paperSize   = ws2.PAPERSIZE_A4
            ws2.page_setup.fitToPage   = True
            ws2.page_setup.fitToWidth  = 1
            ws2.print_options.horizontalCentered = True
            ws2.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75,
                                            header=0.3, footer=0.3)
            ws2.column_dimensions["A"].width = 80

            sheet_w = result["sheet_w"]
            sheet_h = result["sheet_h"]
            sheets_map = {}
            for p in result["placements"]:
                sheets_map.setdefault(p["sheet"], []).append(p)

            # waste_rects levhaya göre
            waste_map_xl = {}
            for wr in result.get("waste_rects", []):
                waste_map_xl.setdefault(wr["sheet"], []).append(wr)

            # Parça numarası etiketleri
            piece_order_xl = {p.get("name",""): i+1 for i,p in enumerate(pieces)}
            name_count_xl  = {}
            for p in result["placements"]:
                base = p["label"].replace(" (D)","")
                name_count_xl[base] = name_count_xl.get(base,0) + 1
            name_seq_xl = {}
            piece_labels_xl = {}
            for p in result["placements"]:
                base = p["label"].replace(" (D)","")
                name_seq_xl[base] = name_seq_xl.get(base,0) + 1
                list_no = piece_order_xl.get(base,"?")
                total   = name_count_xl[base]
                seq     = name_seq_xl[base]
                lbl     = f"{list_no} ({seq}/{total})" if total > 1 else str(list_no)
                key     = f"{p['sheet']}_{p['x']}_{p['y']}"
                piece_labels_xl[key] = lbl

            current_row = 1
            for sh_idx in range(1, result["num_sheets"] + 1):
                # Başlık hücresi
                hc = ws2.cell(row=current_row, column=1,
                              value=f"Levha {sh_idx}/{result['num_sheets']}  -  "
                                    f"{int(sheet_w)}x{int(sheet_h)} mm  -  "
                                    f"Parcalar: {len(sheets_map.get(sh_idx,[]))}")
                hc.fill = S_FILL; hc.font = FH10; hc.alignment = CTR
                ws2.row_dimensions[current_row].height = 18
                current_row += 1

                # Görüntü çiz
                img_buf = draw_sheet_image(
                    sheets_map.get(sh_idx, []), sheet_w, sheet_h,
                    waste_rects=waste_map_xl.get(sh_idx, []),
                    piece_labels=piece_labels_xl,
                    img_w=2400,
                    title=f"Levha {sh_idx}")
                xl_img = XLImage(img_buf)
                # A4 genişliğine sığdır
                xl_img.width  = 520
                xl_img.height = int(520 * sheet_h / sheet_w) + 40
                ws2.add_image(xl_img, f"A{current_row}")

                # Görüntü için satır yüksekliği ayarla
                img_rows = max(1, int(xl_img.height / 15))
                for rr in range(current_row, current_row + img_rows):
                    ws2.row_dimensions[rr].height = 15
                current_row += img_rows + 2

        except Exception as canvas_err:
            import traceback; traceback.print_exc()
            # Canvas hata verse bile Excel'i kaydet

        # ── Kayıt dialog ──
        fname = f"kesim_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        save_path = ask_save_path(fname, [("Excel", "*.xlsx"), ("Tum Dosyalar", "*.*")])
        if not save_path:
            return jsonify({"cancelled": True})

        wb.save(save_path)
        return jsonify({"success": True, "path": save_path,
                        "filename": os.path.basename(save_path)})

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ─── PDF Export ───────────────────────────────────────────────────────────────
def _get_static_dir():
    """EXE veya script modunda static klasörünü bul."""
    candidates = []
    if getattr(sys, 'frozen', False):
        candidates.append(os.path.join(sys._MEIPASS, 'static'))
        candidates.append(os.path.join(os.path.dirname(sys.executable), 'static'))
    candidates.append(os.path.join(os.getcwd(), 'static'))
    candidates.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static'))
    for d in candidates:
        if os.path.isdir(d):
            return d
    return os.path.join(os.getcwd(), 'static')

def _get_font_path(name):
    """Font dosyasını bul — static klasörü öncelikli."""
    static = _get_static_dir()
    p = os.path.join(static, name)
    if os.path.exists(p):
        return p
    # Sistem fontları
    for p in [
        os.path.join(os.environ.get("WINDIR","C:\\Windows"), "Fonts", name),
        f"/usr/share/fonts/truetype/dejavu/{name}",
        f"/usr/share/fonts/truetype/liberation/{name}",
    ]:
        if os.path.exists(p): return p
    return None

@app.route("/api/export_pdf", methods=["POST"])
def api_export_pdf():
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                         Paragraph, Spacer, PageBreak, Image, HRFlowable)
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.graphics.shapes import Drawing, Rect, String, Line
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        return jsonify({"error": "reportlab yuklu degil. pip install reportlab"}), 500

    try:
        # ── Font kayıt (Türkçe) ──
        reg_path  = _get_font_path("DejaVuSans.ttf")
        bold_path = _get_font_path("DejaVuSans-Bold.ttf")
        print(f"[PDF] static dir: {_get_static_dir()}")
        print(f"[PDF] reg font: {reg_path}")
        print(f"[PDF] bold font: {bold_path}")
        if reg_path and bold_path:
            pdfmetrics.registerFont(TTFont("TR",      reg_path))
            pdfmetrics.registerFont(TTFont("TR-Bold", bold_path))
            F_NORMAL = "TR"; F_BOLD = "TR-Bold"
            print("[PDF] DejaVu font yuklendi - Turkce aktif")
        else:
            F_NORMAL = "Helvetica"; F_BOLD = "Helvetica-Bold"
            print("[PDF] UYARI: DejaVu bulunamadi, Helvetica kullaniliyor")

        data    = request.json
        result  = calculate_cost(data)
        pieces  = data.get("pieces", [])
        pricing = data.get("pricing", {})

        fname     = f"kesim_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        save_path = ask_save_path(fname, [("PDF", "*.pdf"), ("Tum Dosyalar", "*.*")])
        if not save_path:
            return jsonify({"cancelled": True})

        doc   = SimpleDocTemplate(save_path, pagesize=A4,
                                  leftMargin=1.5*cm, rightMargin=1.5*cm,
                                  topMargin=1.5*cm,  bottomMargin=1.5*cm)

        PURPLE  = colors.HexColor("#4A1A8A")
        LPURPLE = colors.HexColor("#D4C4F0")
        DARK    = colors.HexColor("#1A0533")
        ORANGE  = colors.HexColor("#E07020")
        WHITE   = colors.white

        def ps(name, **kw):
            defaults = dict(fontName=F_NORMAL, fontSize=10, textColor=DARK, leading=14)
            defaults.update(kw)
            return ParagraphStyle(name, **defaults)

        title_st = ps("t", fontName=F_BOLD,   fontSize=15, textColor=PURPLE, spaceAfter=4)
        sub_st   = ps("s", fontSize=9,         textColor=colors.HexColor("#555"), spaceAfter=8)
        sec_st   = ps("h", fontName=F_BOLD,   fontSize=11, textColor=PURPLE, spaceBefore=8, spaceAfter=4)
        cell_st  = ps("c", fontSize=8)

        story = []

        # ── Logo + Başlık yan yana ──
        logo_path = os.path.join(_get_static_dir(), "logo.jpg")
        if not os.path.exists(logo_path):
            logo_path = None

        # Başlık tablosu — logo solda, metin sağda
        from reportlab.platypus import Table as RLTable, TableStyle as RLTableStyle

        title_para = Paragraph("Atolyemhanem", title_st)
        info_para  = Paragraph(
            f"Kesim Listesi  |  {data.get('material','MDF')}  |  "
            f"{int(data.get('sheet_w',2440))}x{int(data.get('sheet_h',1220))} mm  |  "
            f"Levha: {result['num_sheets']}  |  Fire: %{result['waste_pct']}  |  "
            f"{datetime.now().strftime('%d.%m.%Y %H:%M')}",
            sub_st)

        if logo_path:
            try:
                logo_img = Image(logo_path, width=2.2*cm, height=2.2*cm)
                hdr_tbl = RLTable([[logo_img, [title_para, info_para]]],
                                   colWidths=[2.5*cm, 14*cm])
                hdr_tbl.setStyle(RLTableStyle([
                    ("VALIGN",  (0,0), (-1,-1), "MIDDLE"),
                    ("LEFTPADDING",  (0,0), (0,0), 0),
                    ("RIGHTPADDING", (0,0), (0,0), 8),
                ]))
                story.append(hdr_tbl)
            except Exception:
                story.append(title_para)
                story.append(info_para)
        else:
            story.append(title_para)
            story.append(info_para)

        story.append(HRFlowable(width="100%", thickness=1.5, color=PURPLE, spaceAfter=6))

        # ── Parça Listesi ──
        story.append(Paragraph("Parca Listesi", sec_st))
        bant_map  = {"ust": "UK-1", "alt": "UK-2", "sol": "KK-1", "sag": "KK-2"}
        tdata     = [["#", "Parca Adi", "Boy(cm)", "En(cm)", "Adet", "Bant Kenarlari", "Bant(m)"]]
        for idx, p in enumerate(pieces):
            w_cm = float(p.get("w", 0))
            h_cm = float(p.get("h", 0))
            qty  = int(p.get("qty", 1))
            bant = p.get("bant", [])
            bstr = ", ".join(bant_map.get(s, s) for s in bant) or "-"
            bm   = 0
            for side in bant:
                if side in ["ust", "alt"]: bm += (h_cm * 10 / 1000) * qty  # UK=boy
                else:                       bm += (w_cm * 10 / 1000) * qty  # KK=en
            tdata.append([str(idx+1), p.get("name",""), str(h_cm), str(w_cm), str(qty), bstr, str(round(bm,2))])

        def make_table(data_rows, col_widths, font_n, font_b):
            t = Table(data_rows, colWidths=col_widths, repeatRows=1)
            t.setStyle(TableStyle([
                ("BACKGROUND",   (0,0), (-1,0),  PURPLE),
                ("TEXTCOLOR",    (0,0), (-1,0),  WHITE),
                ("FONTNAME",     (0,0), (-1,0),  font_b),
                ("FONTNAME",     (0,1), (-1,-1), font_n),
                ("FONTSIZE",     (0,0), (-1,-1), 8),
                ("ALIGN",        (0,0), (-1,-1), "CENTER"),
                ("ALIGN",        (1,1), (1,-1),  "LEFT"),
                ("ROWBACKGROUNDS",(0,1),(-1,-1), [LPURPLE, WHITE]),
                ("GRID",         (0,0), (-1,-1), 0.4, colors.HexColor("#A66EE8")),
                ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
                ("TOPPADDING",   (0,0), (-1,-1), 3),
                ("BOTTOMPADDING",(0,0), (-1,-1), 3),
            ]))
            return t

        story.append(make_table(tdata,
            [0.7*cm, 5.2*cm, 1.8*cm, 1.8*cm, 1.2*cm, 3.8*cm, 1.8*cm],
            F_NORMAL, F_BOLD))
        story.append(Spacer(1, 0.4*cm))

        # ── Maliyet Özeti ──
        story.append(Paragraph("Maliyet Ozeti", sec_st))
        cdata = [["Kalem", "Tutar", "Detay"],
            ["Levha Maliyeti", f"{result['cost_sheets']} TL",   f"{result['num_sheets']} levha x {pricing.get('sheet_price',0)} TL"],
            ["Kesim Ucreti",   f"{result['cost_cuts']} TL",     f"{result['num_sheets']} plaka x {pricing.get('cut_price',0)} TL"],
            ["PVC Bant",       f"{result['cost_band']} TL",     f"{result['total_band_m']} m x {pricing.get('band_price',0)} TL"],
            ["Nakliye",        f"{result['cost_shipping']} TL", ""],
            ["TOPLAM",         f"{result['cost_total']} TL",    ""],
        ]
        ct = Table(cdata, colWidths=[4*cm, 3.5*cm, 6*cm])
        ct.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),  (-1,0),  PURPLE),
            ("TEXTCOLOR",     (0,0),  (-1,0),  WHITE),
            ("FONTNAME",      (0,0),  (-1,0),  F_BOLD),
            ("BACKGROUND",    (0,-1), (-1,-1), DARK),
            ("TEXTCOLOR",     (0,-1), (-1,-1), WHITE),
            ("FONTNAME",      (0,-1), (-1,-1), F_BOLD),
            ("FONTNAME",      (0,1),  (-1,-2), F_NORMAL),
            ("FONTSIZE",      (0,0),  (-1,-1), 9),
            ("ROWBACKGROUNDS",(0,1),  (-1,-2), [LPURPLE, WHITE]),
            ("GRID",          (0,0),  (-1,-1), 0.4, colors.HexColor("#A66EE8")),
            ("ALIGN",         (1,0),  (1,-1),  "RIGHT"),
            ("VALIGN",        (0,0),  (-1,-1), "MIDDLE"),
            ("TOPPADDING",    (0,0),  (-1,-1), 4),
            ("BOTTOMPADDING", (0,0),  (-1,-1), 4),
        ]))
        story.append(ct)

        # ── Levha Yerleşim Planları ──
        sheet_w = result["sheet_w"]
        sheet_h = result["sheet_h"]
        sheets_map = {}
        for p in result["placements"]:
            sheets_map.setdefault(p["sheet"], []).append(p)

        # waste_rects levhaya göre grupla
        waste_map = {}
        for wr in result.get("waste_rects", []):
            waste_map.setdefault(wr["sheet"], []).append(wr)

        # Parça numaralandırma — listdeki sıra no ve kaçıncısı
        piece_order = {p.get("name",""): i+1 for i, p in enumerate(pieces)}
        name_count  = {}
        for p in result["placements"]:
            base = p["label"].replace(" (D)","")
            name_count[base] = name_count.get(base, 0) + 1
        name_seq = {}
        placement_labels = {}
        for p in result["placements"]:
            base = p["label"].replace(" (D)","")
            name_seq[base] = name_seq.get(base, 0) + 1
            list_no = piece_order.get(base, "?")
            total   = name_count[base]
            seq     = name_seq[base]
            label   = f"{list_no} ({seq}/{total})" if total > 1 else str(list_no)
            key     = f"{p['sheet']}_{p['x']}_{p['y']}"
            placement_labels[key] = (base, label)

        COLORS_LIST = ["#A8D5BA","#F7C59F","#B5C7E3","#F2A7BB","#C5B4E3",
                       "#A8C7D5","#EDD5A3","#B4D1B0","#D5B4A8","#C7D5A8"]
        pnames    = list({p["label"].replace(" (D)","") for p in result["placements"]})
        color_map = {n: COLORS_LIST[i % len(COLORS_LIST)] for i, n in enumerate(pnames)}

        PW = 16.5 * cm
        PH = 10.0 * cm

        for sh_idx in range(1, result["num_sheets"] + 1):
            story.append(PageBreak())
            story.append(Paragraph(f"Levha {sh_idx} / {result['num_sheets']}", sec_st))
            story.append(Paragraph(
                f"Boyut: {int(sheet_w)} x {int(sheet_h)} mm", sub_st))

            sx = PW / sheet_w
            sy = PH / sheet_h

            d = Drawing(PW, PH)
            # Levha arka planı
            d.add(Rect(0, 0, PW, PH,
                       fillColor=colors.HexColor("#FFFFFF"),
                       strokeColor=PURPLE, strokeWidth=1.5))

            # ── Fire alanları ──
            for wr in waste_map.get(sh_idx, []):
                wx = wr["x"] * sx
                wy = PH - (wr["y"] + wr["h"]) * sy
                ww = wr["w"] * sx
                wh = wr["h"] * sy
                if ww < 1 or wh < 1: continue
                d.add(Rect(wx, wy, ww, wh,
                           fillColor=colors.HexColor("#F5F0E8"),
                           strokeColor=colors.HexColor("#C8A870"),
                           strokeWidth=0.5, strokeDashArray=[2,2]))
                # Fire etiketi
                if ww > 15 and wh > 8:
                    fs = max(4, min(6, ww/10))
                    d.add(String(wx + ww/2, wy + wh/2 + fs*0.6,
                                 "FIRE",
                                 fontSize=fs, textAnchor="middle",
                                 fillColor=colors.HexColor("#A08040")))
                    if ww > 20 and wh > 14:
                        d.add(String(wx + ww/2, wy + wh/2 - fs*0.6,
                                     f"{wr['w_cm']}x{wr['h_cm']}",
                                     fontSize=max(3,fs-1), textAnchor="middle",
                                     fillColor=colors.HexColor("#A08040")))

            # ── Parçalar ──
            for pp in sheets_map.get(sh_idx, []):
                px = pp["x"] * sx
                py = PH - (pp["y"] + pp["h"]) * sy
                pw = pp["w"] * sx
                ph = pp["h"] * sy
                base = pp["label"].replace(" (D)","")
                fc   = colors.HexColor(color_map.get(base, "#C5B4E3"))

                d.add(Rect(px, py, pw, ph,
                           fillColor=fc, strokeColor=PURPLE, strokeWidth=0.5))

                # Bant kenarları
                bant_sides = pp.get("bant_visual") or pp.get("bant") or []
                for side in bant_sides:
                    lw = 1.5
                    if   side == "ust": d.add(Line(px, py+ph, px+pw, py+ph, strokeColor=ORANGE, strokeWidth=lw))
                    elif side == "alt": d.add(Line(px, py,    px+pw, py,    strokeColor=ORANGE, strokeWidth=lw))
                    elif side == "sol": d.add(Line(px, py,    px,    py+ph, strokeColor=ORANGE, strokeWidth=lw))
                    elif side == "sag": d.add(Line(px+pw,py,  px+pw, py+ph, strokeColor=ORANGE, strokeWidth=lw))

                # Ölçüler — parça içinde, üst kenara yakın EN, sol kenara yakın BOY
                key = f"{sh_idx}_{pp['x']}_{pp['y']}"
                pname, plabel = placement_labels.get(key, (base, "?"))
                fs_name = max(6, min(10, pw/8, ph/3))
                fs_dim  = max(5, min(8,  pw/12, ph/5))
                fs_num  = max(5, min(8,  pw/10, ph/4))

                # ReportLab: Y aşağıdan yukarı → py=alt, py+ph=üst
                # Üst içte: EN (w_cm) — üst kenara yakın (py+ph - offset)
                if pw > 15:
                    d.add(String(px + pw/2, py + ph - fs_dim*1.8,
                                 str(pp['w_cm']),
                                 fontSize=fs_dim, textAnchor="middle",
                                 fillColor=colors.HexColor("#3A1800")))

                # Sol içte: BOY (h_cm) — sol kenara yakın, dikey orta
                if ph > 15:
                    d.add(String(px + fs_dim*1.5, py + ph/2,
                                 str(pp['h_cm']),
                                 fontSize=fs_dim, textAnchor="middle",
                                 fillColor=colors.HexColor("#3A1800")))

                # Orta: parça adı + numara
                if pw > 10 and ph > 8:
                    d.add(String(px + pw/2, py + ph/2 + fs_name*0.3,
                                 pname,
                                 fontSize=fs_name, textAnchor="middle",
                                 fillColor=DARK))
                    d.add(String(px + pw/2, py + ph/2 - fs_num*1.2,
                                 plabel,
                                 fontSize=fs_num, textAnchor="middle",
                                 fillColor=colors.HexColor("#4A1A8A")))

            story.append(d)
            story.append(Spacer(1, 0.3*cm))

            # Levha özet tablosu — fire dahil
            lsum = [["#", "Parca", "W x H (cm)", "Konum (cm)"]]
            for pp in sheets_map.get(sh_idx, []):
                key = f"{sh_idx}_{pp['x']}_{pp['y']}"
                _, plabel = placement_labels.get(key, ("", "?"))
                lsum.append([plabel, pp["label"].replace(" (D)",""),
                             f"{pp['w_cm']} x {pp['h_cm']}",
                             f"x={int(pp['x']/10)} y={int(pp['y']/10)}"])
            for i, wr in enumerate(waste_map.get(sh_idx, []), 1):
                lsum.append([f"F{i}", "FIRE",
                             f"{wr['w_cm']} x {wr['h_cm']}", ""])
            story.append(make_table(lsum, [1.2*cm, 4*cm, 3*cm, 3.5*cm], F_NORMAL, F_BOLD))

        doc.build(story)
        return jsonify({"success": True, "path": save_path,
                        "filename": os.path.basename(save_path)})

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ─── Proje Kaydet ─────────────────────────────────────────────────────────────
@app.route("/api/save_project", methods=["POST"])
def api_save_project():
    try:
        data         = request.json
        project_name = data.get("name", "")
        project_data = data.get("data", {})

        # İsim boşsa tkinter ile sor
        if not project_name:
            project_name = ask_string(
                "Proje Kaydet", "Proje adını girin:",
                f"Proje_{datetime.now().strftime('%d-%m-%Y')}"
            )
        if not project_name:
            return jsonify({"cancelled": True})

        fname     = f"{project_name}.json"
        save_path = ask_save_path(fname, [("JSON Proje", "*.json"), ("Tum Dosyalar", "*.*")])
        if not save_path:
            return jsonify({"cancelled": True})

        with open(save_path, "w", encoding="utf-8") as f:
            json.dump(project_data, f, ensure_ascii=False, indent=2)

        return jsonify({"success": True, "path": save_path,
                        "filename": os.path.basename(save_path)})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ─── Proje Yükle ──────────────────────────────────────────────────────────────
@app.route("/api/load_project", methods=["POST"])
def api_load_project():
    try:
        if "file" not in request.files:
            return jsonify({"error": "Dosya bulunamadi"}), 400
        f    = request.files["file"]
        data = json.loads(f.read().decode("utf-8"))
        return jsonify({"success": True, "data": data})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ─── Excel Import ─────────────────────────────────────────────────────────────
@app.route("/api/import_excel", methods=["POST"])
def api_import_excel():
    try:
        import openpyxl
    except ImportError:
        return jsonify({"error": "openpyxl yuklu degil"}), 500
    try:
        if "file" not in request.files:
            return jsonify({"error": "Dosya bulunamadi"}), 400
        wb   = openpyxl.load_workbook(request.files["file"], data_only=True)
        ws   = wb.active
        rows = list(ws.values)

        header_row, col_map = -1, {}
        for i, row in enumerate(rows[:20]):
            rup = [str(c).strip().upper() if c else "" for c in row]
            if any(k in rup for k in {"TANIM","BOY","EN","ADET"}):
                header_row = i
                for j, c in enumerate(rup):
                    if c == "TANIM":          col_map["tanim"] = j
                    elif c == "BOY":          col_map["boy"]   = j
                    elif c == "EN":           col_map["en"]    = j
                    elif c in ("ADET","AD."): col_map["adet"]  = j
                    elif c == "UK1":          col_map["uk1"]   = j
                    elif c == "UK2":          col_map["uk2"]   = j
                    elif c == "KK1":          col_map["kk1"]   = j
                    elif c == "KK2":          col_map["kk2"]   = j
                if "tanim" not in col_map:
                    skip = {"NO","BOY","EN","ADET","AD.","UK1","UK2","KK1","KK2","OLCU","BANT","MM",""}
                    for j, c in enumerate(rup):
                        if j == 0: continue
                        if c not in skip and c: col_map["tanim"] = j; break
                break

        if header_row == -1:
            return jsonify({"error": "Baslik satiri bulunamadi (TANIM, BOY, EN, ADET)"}), 400

        pieces = []
        for row in rows[header_row + 1:]:
            def get(idx):
                if idx is None or idx >= len(row): return ""
                v = row[idx]; return str(v).strip() if v is not None else ""
            def chk(key):
                v = get(col_map.get(key)); return bool(v and v != "0")

            tanim = get(col_map.get("tanim"))
            try: boy = float(get(col_map.get("boy")).replace(",","."))
            except: boy = 0
            try: en  = float(get(col_map.get("en")).replace(",","."))
            except: en  = 0
            try: adet = int(float(get(col_map.get("adet")) or 1))
            except: adet = 1

            if (not tanim and boy == 0 and en == 0) or boy == 0 or en == 0: continue
            pieces.append({"name": tanim or f"Parca {len(pieces)+1}",
                           "h": boy, "w": en, "qty": adet,
                           "uk1": chk("uk1"), "uk2": chk("uk2"),
                           "kk1": chk("kk1"), "kk2": chk("kk2")})

        if not pieces:
            return jsonify({"error": "Gecerli parca bulunamadi"}), 400
        return jsonify({"pieces": pieces, "count": len(pieces)})

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ─── WebView Başlatıcı ────────────────────────────────────────────────────────
def run_flask():
    app.run(host="127.0.0.1", port=5000, debug=False, use_reloader=False)

if __name__ == "__main__":
    flask_thread = threading.Thread(target=run_flask, daemon=True)
    flask_thread.start()
    webview.create_window(
        "Atolyemhanem v2 - Kesim Programi",
        "http://127.0.0.1:5000",
        width=1280, height=900,
        resizable=True, min_size=(300, 300)
    )
    webview.start()
    sys.exit()
